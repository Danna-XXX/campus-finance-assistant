import json
import re
from datetime import datetime
from io import BytesIO

import openpyxl
from openai import OpenAI

from config import SILICONFLOW_BASE_URL, MODEL
from prompts import BILL_ANALYSIS_PROMPT, BILL_CATEGORIZE_PROMPT


CATEGORY_KEYWORDS = {
    "一日三餐": [
        "食堂", "餐厅", "饭", "面", "粥", "外卖", "美团", "饿了么",
        "麦当劳", "肯德基", "汉堡", "火锅", "烧烤", "饺子", "包子",
        "沙县", "兰州", "正餐", "快餐", "小吃", "早点", "早餐", "午餐",
        "晚餐", "夜宵", "盖浇", "煎饼", "炒饭", "炒菜", "烤肉", "烤串",
        "海底捞", "必胜客", "汉堡王", "沙拉", "寿司", "拉面", "螺蛳粉",
        "黄焖鸡", "鸡排", "砂锅", "米饭", "面条", "饮食", "就餐",
    ],
    "偶尔小资": [
        "奶茶", "咖啡", "星巴克", "瑞幸", "冰淇淋", "甜品", "喜茶",
        "茶颜", "霸王茶姬", "贡茶", "古茗", "甜点", "bubble",
        "烘焙", "糕点", "冷饮", "蜜雪", "冰城", "茶饮", "冻", "冰",
    ],
    "日常出行": [
        "滴滴", "地铁", "公交", "打车", "出行", "高铁", "火车", "机票",
        "共享单车", "哈罗", "美团单车", "青桔", "嘀嗒", "顺风车", "快车",
        "出租", "的士", "充值", "交通卡",
    ],
    "娱乐": [
        "电影", "演唱会", "门票", "ktv", "游戏", "视频", "爱奇艺",
        "优酷", "腾讯视频", "bilibili", "steam", "聚餐", "网吧", "剧本",
        "密室", "桌游", "台球", "棋牌", "美甲", "足浴", "健身", "健身房",
        "博物馆", "展览", "演出", "话剧", "livehouse",
    ],
    "学习": [
        "书", "课程", "考试", "培训", "文具", "打印", "教材", "资料",
        "知网", "百度文库", "网课", "复印", "装订", "学习", "图书",
        "论文", "教辅", "笔", "本子", "答题卡",
    ],
    "日用": [
        "超市", "沃尔玛", "华润", "盒马", "便利店", "711", "全家",
        "罗森", "药", "洗漱", "卫生", "生活用品", "京东", "淘宝",
        "物美", "永辉", "大润发", "家乐福", "好邻居", "便民", "生鲜",
        "宜家", "居然", "洗衣", "清洁", "卫生纸", "日化",
    ],
    "快递物流": [
        "顺丰", "圆通", "中通", "申通", "菜鸟", "快递", "物流",
        "韵达", "极兔", "百世", "丰巢",
    ],
}

FAMILY_KEYWORDS = ["爸", "妈", "父", "母", "叔", "姑", "舅", "奶", "爷", "姥", "外婆", "外公", "家人", "爹", "娘"]
RENT_KEYWORDS = ["房东", "中介", "物业", "房租", "租金", "水电", "燃气", "宽带", "网费"]
SKIP_TYPES = ["信用卡还款", "零钱通", "理财通", "基金", "理财", "还款"]


def _is_person_name(name: str) -> bool:
    if len(name) < 2 or len(name) > 4:
        return False
    commercial = ["店", "超", "市", "餐", "厅", "公司", "科技", "网络", "有限", "中心",
                  "服务", "商行", "贸易", "集团", "机构", "平台", "官方", "旗舰"]
    return not any(kw in name for kw in commercial)


def guess_category(counterparty: str, product: str, trade_type: str = "", is_income: bool = False) -> str:
    if is_income:
        text_cp = counterparty.lower()
        if any(kw in text_cp for kw in FAMILY_KEYWORDS):
            return "家人转账"
        return "收入"

    if any(kw in trade_type for kw in SKIP_TYPES):
        return "_skip_"

    text = (counterparty + product).lower()
    cp = counterparty.strip()

    # 快递物流
    for kw in CATEGORY_KEYWORDS["快递物流"]:
        if kw.lower() in text:
            return "快递物流"

    # 转账类：看对方名称判断
    if "转账" in trade_type:
        if any(kw in cp for kw in FAMILY_KEYWORDS):
            return "家人转账"
        if any(kw in cp for kw in RENT_KEYWORDS):
            return "房租水电"
        if _is_person_name(cp):
            return "朋友转账"
        return "其他"

    # 关键词匹配其余类别
    for cat, keywords in CATEGORY_KEYWORDS.items():
        if cat == "快递物流":
            continue
        for kw in keywords:
            if kw.lower() in text:
                return cat

    # 奶茶咖啡单字"茶"匹配
    if "茶" in counterparty and len(counterparty) <= 8:
        return "偶尔小资"

    return "其他"


def categorize_others_with_llm(transactions: list[dict], api_key: str) -> dict[int, str]:
    others = [
        {"idx": i, "交易类型": t.get("trade_type", ""), "交易对方": t["counterparty"],
         "商品说明": t["product"][:30], "金额": t["amount"]}
        for i, t in enumerate(transactions)
        if t.get("category") == "其他" and not t.get("is_income")
    ]
    if not others:
        return {}

    prompt = BILL_CATEGORIZE_PROMPT + json.dumps(others, ensure_ascii=False, indent=2)
    client = OpenAI(api_key=api_key, base_url=SILICONFLOW_BASE_URL)
    try:
        response = client.chat.completions.create(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=2000,
        )
        raw = response.choices[0].message.content
        match = re.search(r"\[.*\]", raw, re.DOTALL)
        if match:
            result_list = json.loads(match.group())
            return {item["idx"]: item["category"] for item in result_list if "idx" in item and "category" in item}
    except Exception:
        pass
    return {}


def parse_wechat_bill(file) -> list[dict]:
    if isinstance(file, bytes):
        wb = openpyxl.load_workbook(BytesIO(file), data_only=True)
    elif isinstance(file, str):
        wb = openpyxl.load_workbook(file, data_only=True)
    else:
        content = file.read()
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(rows):
        if row and row[0] and "交易时间" in str(row[0]):
            header_row_idx = i
            break
    if header_row_idx is None:
        header_row_idx = 16

    transactions = []
    for row in rows[header_row_idx + 1:]:
        if not row or all(cell is None for cell in row):
            continue
        try:
            trade_time = row[0]
            trade_type = str(row[1] or "")
            counterparty = str(row[2] or "")
            product = str(row[3] or "")
            direction = str(row[4] or "")
            amount_raw = row[5]
            status = str(row[7] or "")

            if amount_raw is None:
                continue

            amount = float(str(amount_raw).replace("¥", "").replace(",", "").strip())

            if isinstance(trade_time, datetime):
                time_str = trade_time.strftime("%Y-%m-%d %H:%M")
            else:
                time_str = str(trade_time)[:16]

            is_income = "收入" in direction
            is_expense = "支出" in direction

            if not (is_income or is_expense):
                continue

            category = guess_category(counterparty, product, trade_type, is_income)

            if category == "_skip_":
                continue

            transactions.append({
                "time": time_str,
                "trade_type": trade_type,
                "counterparty": counterparty,
                "product": product,
                "is_income": is_income,
                "amount": amount,
                "status": status,
                "category": category,
            })
        except (ValueError, TypeError, IndexError):
            continue

    return transactions


def aggregate_by_month(transactions: list[dict]) -> dict:
    monthly: dict = {}
    for t in transactions:
        month = t["time"][:7]
        if month not in monthly:
            monthly[month] = {"income": 0.0, "expense": {}, "total_expense": 0.0}
        if t["is_income"]:
            monthly[month]["income"] += t["amount"]
        else:
            cat = t["category"]
            monthly[month]["expense"][cat] = monthly[month]["expense"].get(cat, 0.0) + t["amount"]
            monthly[month]["total_expense"] += t["amount"]
    return dict(sorted(monthly.items()))


def analyze_bill_with_llm(transactions: list[dict], api_key: str) -> dict:
    # Phase 2: LLM batch re-categorize "其他"
    llm_cats = categorize_others_with_llm(transactions, api_key)
    for idx, cat in llm_cats.items():
        if 0 <= idx < len(transactions):
            transactions[idx]["category"] = cat

    monthly_agg = aggregate_by_month(transactions)
    span_months = len(monthly_agg)

    payload = {
        "span_months": span_months,
        "monthly_summary": {
            m: {
                "income": round(d["income"], 2),
                "total_expense": round(d["total_expense"], 2),
                "expense_by_category": {k: round(v, 2) for k, v in d["expense"].items()},
            }
            for m, d in monthly_agg.items()
        },
        "sample_transactions": [
            {
                "time": t["time"],
                "trade_type": t.get("trade_type", ""),
                "counterparty": t["counterparty"],
                "product": t["product"][:30],
                "is_income": t["is_income"],
                "amount": t["amount"],
                "category": t["category"],
            }
            for t in transactions[:30]
        ],
    }

    prompt = BILL_ANALYSIS_PROMPT + json.dumps(payload, ensure_ascii=False, indent=2)

    client = OpenAI(api_key=api_key, base_url=SILICONFLOW_BASE_URL)
    response = client.chat.completions.create(
        model=MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=3000,
        response_format={"type": "json_object"},
    )

    raw = response.choices[0].message.content
    try:
        result = json.loads(raw)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            result = json.loads(match.group())
        else:
            result = {
                "headline": "账单分析完成",
                "full_summary": {},
                "persona": {"type": "未知", "emoji": "🐱", "strengths": [], "blindspots": [], "tips": []},
                "suggested_accounts": {},
            }

    result["monthly_agg"] = monthly_agg
    return result
